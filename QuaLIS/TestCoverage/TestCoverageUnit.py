import datetime
from os import listdir

import keyboard
import time
from configparser import ConfigParser

import openpyxl
import pypyodbc
import xlrd
from selenium.webdriver.common.by import By

from Database import DBUnit
from ObjectRepository import ElementAuditTrail
from TestCoverage import TestCoverageAudittrail
from TestCoverage.TestCoverageAuditTrail import ExpectedAuditTrail
from Utility import BasicOperation, ScreenNavigate, BrowserOperation, ResultFlag, LogOperation, ExceptionHandling, JDBC
from Utility.BrowserOperation import configDriver
from loguru import logger


baseMaster=ConfigParser()
baseMaster.read(BasicOperation.projectDirectory()+"\\ObjectRepository\\ElementBaseMaster.ini")

configDriver=ConfigParser()
configDriver.read(BasicOperation.projectDirectory()+"\\config.ini")

id=configDriver.get("Credential", "loginid")



def unitAdd(driver,name,description,defaultStatus):

    ScreenNavigate.unit(driver,"add")

    countListbefore=driver.find_element(By.XPATH,baseMaster.get("UnitOfMeasurement", "unitCount")).text

    countLIST=countListbefore.split(' ')

    beforeCount=countLIST[4]

    print("before count-->"+str(beforeCount))

    ExceptionHandling.exceptionClick(driver,baseMaster.get("UnitOfMeasurement", "unitAdd"),"Clicked the Unit add button", "Unable to click the unit add button")

    time.sleep(3)
    ExceptionHandling.exceptionSendKeys(driver, baseMaster.get("UnitOfMeasurement", "unitName"), name,"Entered the unit name", "Unable to enter the unit name")
    time.sleep(3)
    ExceptionHandling.exceptionSendKeys(driver, baseMaster.get("UnitOfMeasurement", "unitDescription"), description,"Entered the unit description", "Unable to enter the unit description")
    time.sleep(3)
    if defaultStatus=="Yes":

        defaultStatusToggle=driver.find_element(By.XPATH,baseMaster.get("UnitOfMeasurement", "unitDefaultStatusToggle"))

        if defaultStatusToggle.get_attribute("value")=="false":
            ExceptionHandling.exceptionClick(driver, baseMaster.get("UnitOfMeasurement", "unitDefaultStatus"),"Clicked the default status button","Unable to click the default status")

    esign=""

    ExceptionHandling.exceptionClick(driver,
                              baseMaster.get("UnitOfMeasurement", "unitAddSubmit"),"Clicked the add submit button","unable to click the add submit button")
    time.sleep(3)


    addTime=BasicOperation.timet()

    time.sleep(3)
    element = driver.find_element(By.XPATH, "//span[text()='Base Master']")
    time.sleep(2)
    try:

        driver.execute_script("arguments[0].scrollIntoView();", element)
    except:
        pass

    time.sleep(3)
    element.click()

    output={"addtime":addTime}

    return output


def unitDelete(driver,name,description,defaultStatus):

    ScreenNavigate.unit(driver,"delete")

    try:
        countListbefore = driver.find_element(By.XPATH, baseMaster.get("UnitOfMeasurement", "unitCount")).text

        countLIST = countListbefore.split(' ')

        beforeCount = int(countLIST[len(countLIST) - 1])

        logger.info("Got the count before delete, Count is " + str(beforeCount))

    except:

        try:

            time.sleep(5)

            countListbefore = driver.find_element(By.XPATH, baseMaster.get("UnitOfMeasurement", "unitCount")).text

            countLIST = countListbefore.split(' ')

            beforeCount = int(countLIST[len(countLIST) - 1])

            logger.info("Got the count before delete, Count is "+str(beforeCount))
        except:
            logger.error("Unable to get the count")

    q = driver.find_elements(By.TAG_NAME, "tr")
    qq = len(q)

    # input("enter containertype with description: ")
    unit = name+" "+description+" "+defaultStatus
    for i in range(1, qq):
        unitrow = q[i].text

        if unit.__contains__(name):
            logger.info("Data is available in the grid index".format(unitrow))

            m = str(i)
            delete = "(//span[@data-tip='Delete'])[" + m + "]"
            print(delete)
            try:
                driver.find_element(By.XPATH, delete).click()
                logger.info("Clicked the delete button")
            except:
                logger.error("Unable to click the delete button")

            try:
                driver.find_element(By.XPATH,"//*[text()='OK']").click()
                logger.info("Clicked the delete confirmation button")
            except:
                logger.error("Unable to click the delete confirmation button")

            break

        else:
            logger.error("The delete data is not available in the Grid")


        element = driver.find_element(By.XPATH, "//span[text()='Base Master']")

        try:
            driver.execute_script("arguments[0].scrollIntoView();", element)
        except:
            pass

        try:
            element.click()
            logger.info("Clicked the post base master icon")

        except:
            logger.error("Unable to click the post base master icon")


    time.sleep(2)

    try:

        countListafter = driver.find_element(By.XPATH, baseMaster.get("UnitOfMeasurement", "unitCount")).text

        aftercountLIS = countListafter.split(' ')

        afterCount = int(aftercountLIS[len(aftercountLIS) - 1])

        logger.info("Got the count before delete, Count is " + str(afterCount))

    except:

        try:

            time.sleep(5)

            countListafter = driver.find_element(By.XPATH, baseMaster.get("UnitOfMeasurement", "unitCount")).text

            aftercountLIS = countListafter.split(' ')

            afterCount = int(aftercountLIS[len(aftercountLIS) - 1])

            logger.info("Got the count after delete, Count is " + str(afterCount))
        except:
            logger.error("Unable to get the count")


    if afterCount==beforeCount-1:

        logger.info("Unit delete is working")

        nstatus=JDBC.retunOneValue("select nstatus from unit where sunitname='{}'".format(name))

        if nstatus==-1:
            logger.info("Delete is updated in the database, Now nstatus is -1")
        else:
            logger.error("Delete is not updated in the database, Now nstatus is 1")

    else:
        logger.error("Unit Delete is not working properly, Data is not deleted")

    element = driver.find_element(By.XPATH, "//span[text()='Base Master']")
    time.sleep(2)
    try:
        driver.execute_script("arguments[0].scrollIntoView();", element)
    except:
        pass
    element.click()



def unitEdit(driver,oldName,oldDescription,oldDefaultStatus,newName,newDescription):
    ScreenNavigate.unit(driver,"edit")

    try:
        countListbefore = driver.find_element(By.XPATH, baseMaster.get("UnitOfMeasurement", "unitCount")).text

        countLIST = countListbefore.split(' ')

        beforeCount = countLIST[4]

        logger.info("Got the count before edit the unit detail")

    except:

        try:
            time.sleep(3)

            countListbefore = driver.find_element(By.XPATH, baseMaster.get("UnitOfMeasurement", "unitCount")).text

            countLIST = countListbefore.split(' ')

            beforeCount = countLIST[4]

            logger.info("Got the count before edit the unit detail")

        except:
            logger.error("undefined")

    q = driver.find_elements(By.TAG_NAME, "tr")
    qq = len(q)

    # input("enter containertype with description: ")
    container = oldName + " " + oldDescription + " " + oldDefaultStatus
    for i in range(1, qq):
        unitrow = q[i].text
        print("grid text --> "+unitrow)
        print("old test -->"+oldName)

        if unitrow.__contains__(oldName):

            logger.info("Old file is available at {} index".format(i))
            m = str(i)
            edit = "(//span[@data-tip='Edit'])[" + m + "]"

            try:
                driver.find_element(By.XPATH, edit).click()
                logger.info("Clicked the edit icon")

            except:
                logger.error("Unable to click the edit icon")


            try:
                BasicOperation.clear(driver, baseMaster.get("UnitOfMeasurement", "unitDescription"))

                logger.info("cleared the old unit description")

            except:
                logger.error("unable to clear the old unit description")

            try:
                BasicOperation.clear(driver, baseMaster.get("UnitOfMeasurement", "unitName"))
                logger.info("cleared the old unit name")

            except:
                logger.error("unable to clear the old unit name")

            try:
                BasicOperation.clear(driver, baseMaster.get("UnitOfMeasurement", "unitDescription"))

                logger.info("cleared the old unit description")

            except:
                logger.error("unable to clear the old unit description")

            try:
                BasicOperation.sendKeysXpath(driver, baseMaster.get("UnitOfMeasurement", "unitName"), newName)

                logger.info("Entered the new name")

            except:
                logger.error("unable to enter the new name")



            try:
                BasicOperation.clickXpath(driver,
                                          baseMaster.get("UnitOfMeasurement", "unitEditSubmit"))
                logger.info("Clicked  the add submit button")

            except:
                logger.error("unable to click the add submit button")


            try:

                BasicOperation.clickXpath(driver,"//*[@id='content']/div[1]/nav/span/h2")

                logger.info("Edit popup is closed")

                time.sleep(2)

                try:

                    countListafter = driver.find_element(By.XPATH,
                                                         baseMaster.get("UnitOfMeasurement", "unitCount")).text

                    aftercountLIS = countListafter.split(' ')

                    afterCount = aftercountLIS[len(aftercountLIS) - 1]

                except:
                    time.sleep(3)

                    countListafter = driver.find_element(By.XPATH,
                                                         baseMaster.get("UnitOfMeasurement", "unitCount")).text

                    aftercountLIS = countListafter.split(' ')

                    afterCount = aftercountLIS[len(aftercountLIS) - 1]

                if beforeCount == afterCount:
                    logger.info("New data is not created")
                    logger.info("Edit is working properly to the count")
                elif afterCount > beforeCount:
                    logger.info("New data is created for the edit action")

                element = driver.find_element(By.XPATH, "//span[text()='Base Master']")

                try:
                    driver.execute_script("arguments[0].scrollIntoView();", element)
                except:
                    pass

                element.click()



            except:

                logger.error("Edit popup is not closed")


            break

        else:
            logger.error("The old name is not available in the grid")


def auditTrailUnitAdd(driver,name,description,defaultStatus):

    esign=""

    beforeCount=TestCoverageAudittrail.auditTrailRecordCount(driver)

    query = "select sfirstname from users where sloginid='{}'".format(id)

    query2 = "select slastname from users where sloginid='{}'".format(id)

    userName=JDBC.retunOneValue(query)+" "+JDBC.retunOneValue(query2)

    time.sleep(5)

    output = unitAdd(driver, name, description, "No")

    configDriver = ConfigParser()

    configDriver.read(BasicOperation.projectDirectory() + "\\config.ini")

    userRole=configDriver.get("Credential","userRole")

    auditTrailRecord={"AuditDateAndTime":output.get("addtime"),"AuditAction":"ADD UNIT","comments":"Unit Name: {};Description: {};Default Status: No;".format(name,description),"userName":"{}".format(userName),"userRole":"{}".format(userRole),"ActionType":"SYSTEM","ModuleName":"Base Master","FormName":"Unit of Measurement","esignComments":"{}".format(esign)}

    expectedAuditAction= auditTrailRecord.get("AuditAction")

    expectedComments=auditTrailRecord.get("comments")

    expectedUserName=auditTrailRecord.get("userName")

    expectedUserRole=auditTrailRecord.get("userRole")

    expectedActionType = auditTrailRecord.get("ActionType")

    expectedModuleName  = auditTrailRecord.get("ModuleName")

    expectedFormName = auditTrailRecord.get("FormName")

    expectedEsignComments = auditTrailRecord.get("esignComments")

    expectedAuditTrail={}

    expectedAuditTrail["expectedAuditAction"]=expectedAuditAction

    expectedAuditTrail["expectedEsignComments"]=expectedEsignComments

    expectedAuditTrail["expectedFormName"]=expectedFormName

    expectedAuditTrail["expectedModuleName"]=expectedModuleName

    expectedAuditTrail["expectedActionType"]=expectedActionType

    expectedAuditTrail["expectedUserRole"]=expectedUserRole

    expectedAuditTrail["expectedUserName"]=expectedUserName

    expectedAuditTrail["expectedComments"]=expectedComments

    time.sleep(3)

    afterCount=TestCoverageAudittrail.auditTrailRecordCount(driver)

    logger.info("After count is "+str(afterCount))

    auditActionList=driver.find_elements(By.XPATH, "//tbody[@role='presentation']/tr[2]/td[2]")

    if afterCount==beforeCount:
        ResultCase1="FAIL"
        result="Audit trail is not captured"
        logger.error(result)

    elif afterCount==beforeCount+1:

        i=2

        ResultCase1 = "PASS"

        result="Audit trail is captured"

        logger.info(result)


        auditTrail(driver,expectedAuditTrail)



def downloadPDF(driver):

    time.sleep(2)

    download = ConfigParser()
    download.read(BasicOperation.projectDirectory() + "\\config.ini")

    downloadPath=download.get("config", "download")

    onlyfiles = [f for f in listdir(downloadPath)]

    beforeCount=len(onlyfiles)

    beforeCountResult="Number of folder and files before download "+str(beforeCount)

    logger.info(beforeCountResult)

    ScreenNavigate.unit(driver,"downloadPDF")

    try:
        BasicOperation.clickXpath(driver,   baseMaster.get("UnitOfMeasurement", "unitDownloadPDF"))
        result="Clicked the Download PDF button"

        LogOperation.logInfo(result)

        logger.info(result)
    except Exception as e:

        result="Unable to click the Download PDF button, It causes the exception, Exception Detail---> " + str(e)

        LogOperation.logError(result)

        logger.error(result)

    time.sleep(5)

    onlyfiles2 = [f for f in listdir(downloadPath)]

    afterCount = len(onlyfiles2)

    afterCountResult = "Number of folder and files after download " + str(afterCount)

    logger.info(afterCountResult)

    if afterCount == beforeCount + 1:
        logger.info("File is downloaded")

    elif afterCount == beforeCount:
        logger.error("File is not downloaded")

    else:
        logger.error("File is not downloaded properly")


    try:
        element = driver.find_element(By.XPATH, "//span[text()='Base Master']")

        try:
            driver.execute_script("arguments[0].scrollIntoView();", element)
        except:
            pass

        element.click()

        result="Clicked the basemaster icon"

        LogOperation.logInfo(result)

        logger.info(result)

    except Exception as e:

        result="Unable to click the basemaster icon -- post condition, It causes exception-->" + str(e)
        LogOperation.logError(result)

        logger.error(result)




def downloadExcel(driver):

    download = ConfigParser()
    download.read(BasicOperation.projectDirectory() + "\\config.ini")

    downloadPath=download.get("config", "download")

    oldFileList = [f for f in listdir(downloadPath)]

    beforeCount=len(oldFileList)

    beforeCountResult="Number of folder and files before download "+str(beforeCount)

    logger.info(beforeCountResult)

    ScreenNavigate.unit(driver,"downloadExcel")

    time.sleep(1)

    try:
        BasicOperation.clickXpath(driver,   baseMaster.get("UnitOfMeasurement", "unitDownloadExcel"))

        Result="Clicked the Download Excel button"

        LogOperation.logInfo(Result)

        logger.info(Result)

    except Exception as e:
        result="Unable to click the Download excel button, It causes the exception, Exception Detail---> "+str(e)
        LogOperation.logError(result)
        logger.error(result)
        time.sleep(2)

    for i in range(0, 20):

        onlyfiles2 = [f for f in listdir(downloadPath)]

        afterCount = len(onlyfiles2)

        if afterCount == beforeCount + 1:

            afterCountResult = "Number of folder and files after download " + str(afterCount)

            logger.info(afterCountResult)

            logger.info("File is downloaded after " + str(i/2) + " seconds")

            list2 = onlyfiles2

            list1 = oldFileList

            oldFile=""

            for i in oldFileList:
                oldFile=oldFile+i

            newFile=None

            for i in onlyfiles2:
                if oldFile.__contains__(i):
                   pass
                else:
                    newFile=i
                    logger.info("The new downloaded file name is "+i)

            if newFile==None:
                pass
            else:
                file=downloadPath+newFile

                open_workbook = openpyxl.load_workbook(file)
                open_worksheet = open_workbook["Sheet1"]

                nameHeader=open_worksheet.cell(row = 1, column = 1).value

                descriptionHeader=open_worksheet.cell(row = 1, column = 2).value

                defaultStatusHeader=open_worksheet.cell(row = 1, column = 3).value

                if nameHeader=="Unit Name":
                    logger.info("The unit name header is displayed properly")
                else:
                    logger.error("The unit name header is not displayed properly")

                if descriptionHeader=="Description":
                    logger.info("The unit description header is displayed properly")
                else:
                    logger.error("The unit name header is not displayed properly")

                if defaultStatusHeader=="Default Status":
                    logger.info("The unit default status header is displaying properly")
                else:
                    logger.error("The unit default status header is not displaying properly")


                excelUnitCount = open_worksheet.max_row-1

                dbUnitCount=JDBC.unitCount()

                if excelUnitCount==dbUnitCount:
                    logger.info("The number of data is matching with Excel")

                    dbdata=JDBC.unitCount()

                else:
                    logger.error("The number of data is not matching with Excel")

                logger.info("db"+str(dbUnitCount))

            loc=downloadPath+newFile

            wb = openpyxl.load_workbook(loc)
            sh = wb["Sheet1"]

            total = []
            for i in range(2, sh.max_row + 1):
                name = sh.cell(row=i, column=1).value
                description = sh.cell(row=i, column=2).value
                defaultValue = sh.cell(row=i, column=3).value
                if defaultValue == "No":
                    defaultValue = "4"
                elif defaultValue == "Yes":
                    defaultValue = "3"
                else:
                    defaultValue = "undefined"

                data = name + description + defaultValue

                total.append(data)

            DBUnit.unitExportEXCELDBCompare(total)

            break


        else:
            logger.info("File is not downloaded till {} seconds".format(i/2))

        time.sleep(0.5)



    try:
            element = driver.find_element(By.XPATH, "//span[text()='Base Master']")

            time.sleep(2)
            try:
                driver.execute_script("arguments[0].scrollIntoView();", element)
            except:
                pass

            element.click()

            result="Clicked the basemaster icon"

            logger.info(result)

            LogOperation.logInfo(result)

    except Exception as e:

            result="Unable to click the basemaster icon -- post condition, It causes exception-->"+str(e)

            logger.error(result)

            LogOperation.logError(result)


def unitFilterContains(driver):
    ScreenNavigate.unit(driver,"Filter")

    BasicOperation.clickXpath(driver, baseMaster.get("UnitOfMeasurement", "unitNameFilter"))

    try:
        unitNameFilter = driver.find_element(By.XPATH, "(//span[@role='listbox'])[2]")
        logger.info("clicked the filter icon")

    except:
        logger.error("Unable to click the filter icon")

    try:
        unitNameFilter.click()
        logger.info("Clicked filter drop down")
    except:
        logger.error("unable to click filter drop down")


    try:
        driver.find_element(By.XPATH, "//*[text()='Contains']").click()
        logger.info("selected contains")
    except:
        pass


    textA=" (//input[@class='k-textbox'])[1]"

    clear="//button[@class='k-button' and text()='Clear']"

    filter="//button[@class='k-button k-primary' and text()='Filter']"

    name="cm"
    try:
        BasicOperation.sendKeysXpath(driver, textA, name)
        logger.info("Entered name value")
    except:
        logger.error("Unable to enter name value")


    try:
        BasicOperation.clickXpath(driver, filter)
        logger.info("Clicked filter find buttonj")
    except:
        logger.error("Unable to click filter find button")





    rowList = driver.find_elements(By.TAG_NAME, "tr")

    for i in rowList:
        rowText=i.text

        if rowText.__contains__(name):
            print("data shows properly")





def auditTrailUnitEdit(driver,oldName,oldDescription,defaultStatus,newName,newDescription):

    esign=""

    beforeCount=TestCoverageAudittrail.auditTrailRecordCount(driver)

    query = "select sfirstname from users where sloginid='{}'".format(id)

    query2 = "select sfirstname from users where sloginid='{}'".format(id)

    userName=JDBC.retunOneValue(query)+" "+JDBC.retunOneValue(query2)

    unitEdit(driver, oldName, oldDescription, "No",newName,newDescription)

    configDriver = ConfigParser()

    configDriver.read(BasicOperation.projectDirectory() + "\\config.ini")

    userRole=configDriver.get("Credential","userRole")


    auditTrailRecord={"AuditAction":"EDIT UNIT","comments":"Unit Name: {}-> {};".format(oldName,newName),"userName":"{}".format(userName),"userRole":"{}".format(userRole),"ActionType":"SYSTEM","ModuleName":"Base Master","FormName":"Unit of Measurement","esignComments":"".format(esign)}

    expectedAuditAction= auditTrailRecord.get("AuditAction")

    expectedComments=auditTrailRecord.get("comments")

    expectedUserName=auditTrailRecord.get("userName")

    expectedUserRole=auditTrailRecord.get("userRole")

    expectedActionType = auditTrailRecord.get("ActionType")

    expectedModuleName  = auditTrailRecord.get("ModuleName")

    expectedFormName = auditTrailRecord.get("FormName")

    expectedEsignComments = auditTrailRecord.get("esignComments")

    expectedAuditTrail={}

    expectedAuditTrail["expectedAuditAction"]=expectedAuditAction

    expectedAuditTrail["expectedEsignComments"]=expectedEsignComments

    expectedAuditTrail["expectedFormName"]=expectedFormName

    expectedAuditTrail["expectedModuleName"]=expectedModuleName

    expectedAuditTrail["expectedActionType"]=expectedActionType

    expectedAuditTrail["expectedUserRole"]=expectedUserRole

    expectedAuditTrail["expectedUserName"]=expectedUserName

    expectedAuditTrail["expectedComments"]=expectedComments

    time.sleep(3)

    afterCount=TestCoverageAudittrail.auditTrailRecordCount(driver)

    logger.info("After count is "+str(afterCount))

    auditActionList=driver.find_elements(By.XPATH, "//tbody[@role='presentation']/tr[2]/td[2]")

    if afterCount==beforeCount:
        ResultCase1="FAIL"
        result="Audit trail is not captured"
        logger.error(result)

    elif afterCount==beforeCount+1:

        i=2

        ResultCase1 = "PASS"

        result="Audit trail is captured"

        logger.info(result)


        auditTrail(driver,expectedAuditTrail)

def auditTrailUnitDelete(driver,name,description,defaultStatus):

    esign=""

    beforeCount=TestCoverageAudittrail.auditTrailRecordCount(driver)

    query = "select sfirstname from users where sloginid='{}'".format(id)

    query2 = "select sfirstname from users where sloginid='{}'".format(id)

    userName=JDBC.retunOneValue(query)+" "+JDBC.retunOneValue(query2)

    unitDelete(driver,name,description,"yes")

    configDriver = ConfigParser()

    configDriver.read(BasicOperation.projectDirectory() + "\\config.ini")

    userRole=configDriver.get("Credential","userRole")

    auditTrailRecord = {"AuditAction": "DELETE UNIT",
                        "comments": "Unit Name: {};Description: {};Default Status: No;	".format(name, description),
                        "userName": "{}".format(userName), "userRole": "{}".format(userRole), "ActionType": "SYSTEM",
                        "ModuleName": "Base Master", "FormName": "Unit of Measurement", "esignComments": ""}

    expectedAuditAction= auditTrailRecord.get("AuditAction")

    expectedComments=auditTrailRecord.get("comments")

    expectedUserName=auditTrailRecord.get("userName")

    expectedUserRole=auditTrailRecord.get("userRole")

    expectedActionType = auditTrailRecord.get("ActionType")

    expectedModuleName  = auditTrailRecord.get("ModuleName")

    expectedFormName = auditTrailRecord.get("FormName")

    expectedEsignComments = auditTrailRecord.get("esignComments")

    expectedAuditTrail={}

    expectedAuditTrail["expectedAuditAction"]=expectedAuditAction

    expectedAuditTrail["expectedEsignComments"]=expectedEsignComments

    expectedAuditTrail["expectedFormName"]=expectedFormName

    expectedAuditTrail["expectedModuleName"]=expectedModuleName

    expectedAuditTrail["expectedActionType"]=expectedActionType

    expectedAuditTrail["expectedUserRole"]=expectedUserRole

    expectedAuditTrail["expectedUserName"]=expectedUserName

    expectedAuditTrail["expectedComments"]=expectedComments

    time.sleep(3)

    afterCount=TestCoverageAudittrail.auditTrailRecordCount(driver)

    logger.info("After count is "+str(afterCount))

    auditActionList=driver.find_elements(By.XPATH, "//tbody[@role='presentation']/tr[2]/td[2]")

    if afterCount==beforeCount:
        ResultCase1="FAIL"
        result="Audit trail is not captured"
        logger.error(result)

    elif afterCount==beforeCount+1:

        i=2

        ResultCase1 = "PASS"

        result="Audit trail is captured"

        logger.info(result)


        auditTrail(driver,expectedAuditTrail)
def unitDuplicate(driver):
    # TestCoverageUnit.unitAdd(driver,"name","d","yes")

    ScreenNavigate.unit(driver)

    countListbefore = driver.find_element(By.XPATH, baseMaster.get("UnitOfMeasurement", "unitCount")).text

    countLIST = countListbefore.split(' ')

    beforeCount = countLIST[4]

    print("before count-->" + str(beforeCount))

    ExceptionHandling.exceptionClick(driver, baseMaster.get("UnitOfMeasurement", "unitAdd"),
                                     "Clicked the Unit add button", "Unable to click the unit add button")

    ExceptionHandling.exceptionSendKeys(driver, baseMaster.get("UnitOfMeasurement", "unitName"), "name",
                                        "Entered the unit name", "Unable to enter the unit name")

    ExceptionHandling.exceptionClick(driver,
                                     baseMaster.get("UnitOfMeasurement", "unitAddSubmit"),
                                     "Clicked the add submit button", "unable to click the add submit button")

    time.sleep(2)

    try:
        BasicOperation.clickXpath(driver, "//*[@id='content']/div[1]/nav/span/h2")

        logger.error("It accept duplicate records")

    except:

        try:
            BasicOperation.clickXpath(driver, "//*[@id='add-user']")

            logger.info("it not accept duplicate records")

            for i in range(0, 10):
                try:
                    BasicOperation.clickXpath(driver, "//*[text()='Cancel']")
                    logger.info("Clicked the cancel button to close the popup")
                    break
                except:
                    time.sleep(1)

        except:

            logger.error("Undefined")

            pass

unique="unique"

def unitDuplicate(driver):

    unitAdd(driver,unique,"d","yes")

    ScreenNavigate.unit(driver,"Duplicate")

    countListbefore=driver.find_element(By.XPATH,baseMaster.get("UnitOfMeasurement", "unitCount")).text

    countLIST=countListbefore.split(' ')

    beforeCount=countLIST[4]

    print("before count-->"+str(beforeCount))

    ExceptionHandling.exceptionClick(driver,baseMaster.get("UnitOfMeasurement", "unitAdd"),"Clicked the Unit add button", "Unable to click the unit add button")

    ExceptionHandling.exceptionSendKeys(driver, baseMaster.get("UnitOfMeasurement", "unitName"), unique,"Entered the unit name", "Unable to enter the unit name")

    ExceptionHandling.exceptionClick(driver,
                              baseMaster.get("UnitOfMeasurement", "unitAddSubmit"),"Clicked the add submit button","unable to click the add submit button")

    time.sleep(2)

    try:
        BasicOperation.clickXpath(driver,"//*[@id='content']/div[1]/nav/span/h2")

        logger.error("It accept duplicate records")

        BasicOperation.screenshot(driver,
                                  BasicOperation.projectDirectory() + "Report\\Screenshot\\Unit\\UnitAddDuplicate\\duplicateEntry.error.png")


    except:

        try:
            BasicOperation.clickXpath(driver, "//*[@id='add-user']")

            infoScreenshot=BasicOperation.projectDirectory()+"\\Report\\Screenshot\\Unit\\UnitAddDuplicate\\duplicateEntry.info.png"

            print(infoScreenshot)

            time.sleep(2)

            BasicOperation.screenshot(driver,infoScreenshot)

            logger.info("it not accept duplicate records")

            for i in range(0,10):
                try:
                    BasicOperation.clickXpath(driver,"//*[text()='Cancel']")
                    logger.info("Clicked the cancel button to close the popup")
                    break
                except:
                    time.sleep(1)

        except:

            logger.error("Undefined")

            pass






def unitFilterEquals(driver):
    ScreenNavigate.unit(driver,"Filter")

    BasicOperation.clickXpath(driver, baseMaster.get("UnitOfMeasurement", "unitNameFilter"))

    try:
        unitNameFilter = driver.find_element(By.XPATH, "(//span[@role='listbox'])[2]")
        logger.info("clicked the filter icon")

    except:
        logger.error("Unable to click the filter icon")

    try:
        unitNameFilter.click()
        logger.info("Clicked filter drop down")
    except:
        logger.error("unable to click filter drop down")


    try:
        driver.find_element(By.XPATH, "//*[text()='Is equal to']").click()
        logger.info("selected contains")
    except:
        pass


    textA=" (//input[@class='k-textbox'])[1]"

    clear="//button[@class='k-button' and text()='Clear']"

    filter="//button[@class='k-button k-primary' and text()='Filter']"

    name="cm"
    try:
        BasicOperation.sendKeysXpath(driver, textA, name)
        logger.info("Entered name value")
    except:
        logger.error("Unable to enter name value")


    try:
        BasicOperation.clickXpath(driver, filter)
        logger.info("Clicked filter find buttonj")
    except:
        logger.error("Unable to click filter find button")


    rowList = driver.find_elements(By.TAG_NAME, "tr")

    for i in rowList:
        rowText=i.text

        if rowText==name:
            logger.info("Equal Filter is working properly")
        else:
            logger.error("Equal Filter is not working properly")





def unitFilter(driver):
    ScreenNavigate.unit(driver,"Filter")

    BasicOperation.clickXpath(driver, baseMaster.get("UnitOfMeasurement", "unitNameFilter"))

    try:
        unitNameFilter = driver.find_element(By.XPATH, "(//span[@role='listbox'])[2]")
        logger.info("clicked the filter icon")

    except:
        logger.error("Unable to click the filter icon")

    try:
        unitNameFilter.click()
        logger.info("Clicked filter drop down")
    except:
        logger.error("unable to click filter drop down")


    try:
        driver.find_element(By.XPATH, "//*[text()='Contains']").click()
        logger.info("selected contains")
    except:
        pass


    textA=" (//input[@class='k-textbox'])[1]"

    clear="//button[@class='k-button' and text()='Clear']"

    filter="//button[@class='k-button k-primary' and text()='Filter']"

    name="cm"
    try:
        BasicOperation.sendKeysXpath(driver, textA, name)
        logger.info("Entered name value")
    except:
        logger.error("Unable to enter name value")


    try:
        BasicOperation.clickXpath(driver, filter)
        logger.info("Clicked filter find buttonj")
    except:
        logger.error("Unable to click filter find button")

    rowList = driver.find_elements(By.TAG_NAME, "tr")

    for i in rowList:
        rowText=i.text

        if rowText.__contains__(name):
            print("data shows properly")




def auditTrail(driver,expectedAuditTrailResult):

    expectedAuditAction=expectedAuditTrailResult.get("expectedAuditAction")


    expectedEsignComments=expectedAuditTrailResult.get("expectedEsignComments")

    expectedFormName = expectedAuditTrailResult.get("expectedFormName")


    expectedModuleName = expectedAuditTrailResult.get("expectedModuleName")


    expectedActionType = expectedAuditTrailResult.get("expectedActionType")


    expectedUserRole = expectedAuditTrailResult.get("expectedUserRole")



    expectedUserName = expectedAuditTrailResult.get("expectedUserName")


    expectedComments = expectedAuditTrailResult.get("expectedComments")



    actualAuditActionUI = driver.find_element(By.XPATH, "//tbody[@role='presentation']/tr[2]/td[3]").text

    if actualAuditActionUI == expectedAuditAction:

        logger.info("Audit action is properly mentioned in UI")
        logger.info("Expected--->" + expectedAuditAction)
        logger.info("Actual  --->" + actualAuditActionUI)

    else:
        logger.error("Audit action is not properly mentioned in UI")
        logger.error("Expected--->" + expectedAuditAction)
        logger.error("Actual  --->" + actualAuditActionUI)

    actualAuditActionDB = JDBC.retunOneValue(
        "select sauditaction from auditaction where nauditcode=(select COUNT(*) from auditaction )")

    if actualAuditActionDB == expectedAuditAction:

        logger.info("Audit action is properly mentioned in db")
        logger.info("Expected--->" + expectedAuditAction)
        logger.info("Actual  --->" + actualAuditActionDB)
    else:
        logger.error("Audit action is not properly mentioned db")
        logger.error("Expected--->" + expectedAuditAction)
        logger.error("Actual  --->" + actualAuditActionDB)

    if actualAuditActionUI == actualAuditActionDB:
        logger.info("Audit action is same in the db and ui")
        logger.info("actualAuditActionUI--->" + actualAuditActionUI)
        logger.info("actualAuditActionDB  --->" + actualAuditActionDB)
    else:
        logger.error("Audit action is not same in the db and ui")
        logger.error("actualAuditActionUI--->" + actualAuditActionUI)
        logger.error("actualAuditActionDB  --->" + actualAuditActionDB)

    actualCommentsUI = driver.find_element(By.XPATH, "//tbody[@role='presentation']/tr[2]/td[4]").text

    if actualCommentsUI == expectedComments:
        logger.info("Audit comment is properly mentioned in UI")
        logger.info("Expected--->" + expectedComments)
        logger.info("Actual  --->" + actualCommentsUI)
    else:
        logger.error("Audit comment is not properly mentioned in UI")
        logger.error("Expected--->" + expectedComments)
        logger.error("Actual  --->" + actualCommentsUI)

    actualCommentDb = JDBC.retunOneValue(
        "select scomments from auditcomments where nauditcode=(select COUNT(*)from auditaction)")

    if actualCommentDb == expectedComments:

        logger.info("Audit comment is properly mentioned in db")
        logger.info("Expected--->" + expectedComments)
        logger.info("Actual  --->" + actualCommentDb)
    else:
        logger.error("Audit action is not properly mentioned db")
        logger.error("Expected--->" + expectedComments)
        logger.error("Actual  --->" + actualCommentDb)

    if actualCommentsUI == actualCommentDb:
        logger.info("user role is same in the db and ui")
        logger.info("actualCommentsUI--->" + actualCommentsUI)
        logger.info("actualCommentDb  --->" + actualCommentDb)
    else:
        logger.error("user role is not same in the db and ui")
        logger.error("actualCommentsUI--->" + actualCommentsUI)
        logger.error("actualCommentDb  --->" + actualCommentDb)

    actualUserNameUI = driver.find_element(By.XPATH, "//tbody[@role='presentation']/tr[2]/td[5]").text

    if actualUserNameUI == expectedUserName:
        logger.info("user name is properly mentioned in UI")
        logger.info("Expected--->" + expectedUserName)
        logger.info("Actual  --->" + actualUserNameUI)
    else:
        logger.error("user name is not properly mentioned in UI")
        logger.error("Expected--->" + expectedUserName)
        logger.error("Actual  --->" + actualUserNameUI)

    actualUserRoleUI = driver.find_element(By.XPATH, "//tbody[@role='presentation']/tr[2]/td[6]").text

    if actualUserRoleUI == expectedUserRole:
        logger.info("user role is properly mentioned in UI")
        logger.info("Expected--->" + expectedUserRole)
        logger.info("Actual  --->" + actualUserRoleUI)

    else:
        logger.error("Audit action is not properly mentioned in UI")
        logger.error("Expected--->" + expectedUserRole)
        logger.error("Actual  --->" + actualUserRoleUI)

    queryr = "select suserrolename from userrole where nuserrolecode=(select nuserrole from auditaction where nauditcode=(select COUNT(*) from auditaction ))"

    actualUserRoleDB = JDBC.retunOneValue(queryr)

    if actualUserRoleDB == expectedUserRole:
        logger.info("user role is properly mentioned in db")
        logger.info("Expected--->" + expectedUserRole)
        logger.info("Actual  --->" + actualUserRoleDB)


    else:
        logger.error("user role is not properly mentioned db")
        logger.info("Expected--->" + expectedUserRole)
        logger.info("Actual  --->" + actualUserRoleDB)

    if actualUserRoleUI == actualUserRoleDB:
        logger.info("user role is same in the db and ui")
        logger.info("actualUserRoleUI--->" + actualUserRoleUI)
        logger.info("actualUserRoleDB  --->" + actualUserRoleDB)

    else:
        logger.error("user role is not same in the db and ui")
        logger.error("actualUserRoleUI--->" + actualUserRoleUI)
        logger.error("actualUserRoleDB  --->" + actualUserRoleDB)

    actualActionTypeUI = driver.find_element(By.XPATH, "//tbody[@role='presentation']/tr[2]/td[7]").text

    if actualActionTypeUI == expectedActionType:
        logger.info("action type is properly mentioned in UI")
        logger.info("Expected--->" + expectedActionType)
        logger.info("Actual  --->" + actualActionTypeUI)

    else:
        logger.error("action type is not properly mentioned in UI")
        logger.info("Expected--->" + expectedActionType)
        logger.info("Actual  --->" + actualActionTypeUI)

    actionTypeQuery = "select sactiontype from auditaction where nauditcode=(select COUNT(*) from auditaction )"

    actualActionTypeDB = JDBC.retunOneValue(actionTypeQuery)

    if actualActionTypeDB == expectedActionType:
        logger.info("action type is properly mentioned in db")
        logger.info("Expected--->" + expectedActionType)
        logger.info("Actual  --->" + actualActionTypeDB)

    else:
        logger.error("action type is not properly mentioned db")
        logger.error("Expected--->" + expectedActionType)
        logger.error("Actual  --->" + actualActionTypeDB)

    if actualActionTypeDB == actualActionTypeUI:
        logger.info("action type is same in the db and ui")
        logger.info("actualActionTypeDB--->" + actualActionTypeDB)
        logger.info("actualActionTypeUI  --->" + actualActionTypeUI)
    else:
        logger.error("action type is not same in the db and ui")
        logger.error("actualActionTypeDB--->" + actualActionTypeDB)
        logger.error("actualActionTypeUI  --->" + actualActionTypeUI)

    actualModuleNameUI = driver.find_element(By.XPATH, "//tbody[@role='presentation']/tr[2]/td[8]").text

    if actualModuleNameUI == expectedModuleName:
        logger.info("Expected--->" + expectedModuleName)
        logger.info("Actual  --->" + actualModuleNameUI)
        logger.info("Module name is properly mentioned in UI")

    else:
        logger.error("Module name  is not properly mentioned in UI")
        logger.error("Expected--->" + expectedModuleName)
        logger.error("Actual  --->" + actualModuleNameUI)

    moduleNameQuery = "select smodulename from qualismodule where nmodulecode=(select nmodulecode from auditaction where nauditcode=(select COUNT(*) from auditaction ))"

    actualModuleNameDB = JDBC.retunOneValue(moduleNameQuery)

    if actualModuleNameDB == expectedModuleName:

        logger.info("module is properly mentioned in db")
        logger.info("Expected--->" + expectedActionType)
        logger.info("Actual  --->" + actualModuleNameDB)
    else:
        logger.error("module is not properly mentioned db")
        logger.error("Expected--->" + expectedActionType)
        logger.error("Actual  --->" + actualModuleNameDB)

    if actualModuleNameDB == actualModuleNameUI:
        logger.info("module is same in the db and ui")
        logger.info("actualModuleNameDB--->" + actualModuleNameDB)
        logger.info("actualModuleNameUI  --->" + actualModuleNameUI)
    else:
        logger.error("module is not same in the db and ui")
        logger.error("actualModuleNameDB--->" + actualModuleNameDB)
        logger.error("actualModuleNameUI  --->" + actualModuleNameUI)

    actualFormNameUI = driver.find_element(By.XPATH, "//tbody[@role='presentation']/tr[2]/td[9]").text

    if actualFormNameUI == expectedFormName:
        logger.info("form name is properly mentioned in UI")
        logger.info("Expected--->" + expectedModuleName)
        logger.info("Actual  --->" + actualFormNameUI)

    else:
        logger.error("form name  is not properly mentioned in UI")
        logger.error("Expected--->" + expectedModuleName)
        logger.error("Actual  --->" + actualFormNameUI)

    formNameQuery = "select sformname from qualisforms where nformcode=(select nformcode from auditaction where nauditcode=(select COUNT(*) from auditaction )) and nformcode<>-2"
    actualFormNameDB = JDBC.retunOneValue(formNameQuery)

    if actualFormNameDB == expectedFormName:
        logger.info("form is properly mentioned in db")

        logger.info("Expected--->" + expectedFormName)
        logger.info("Actual  --->" + actualFormNameDB)

    else:
        logger.error("form is not properly mentioned db")

        logger.error("Expected--->" + expectedFormName)
        logger.error("Actual  --->" + actualFormNameDB)

    if actualFormNameUI == actualFormNameDB:
        logger.info("form is same in the db and ui")
        logger.info("actualFormNameUI--->" + actualFormNameUI)
        logger.info("actualFormNameDB  --->" + actualFormNameDB)
    else:
        logger.error("form is not same in the db and ui")
        logger.error("actualFormNameUI--->" + actualFormNameUI)
        logger.error("actualFormNameDB  --->" + actualFormNameDB)

    actualEsignCommentsUI = driver.find_element(By.XPATH, "//tbody[@role='presentation']/tr[2]/td[10]").text

    if actualEsignCommentsUI == expectedEsignComments:
        logger.info("Esign comment name is properly mentioned in UI")
        logger.info("Expected--->" + expectedEsignComments)
        logger.info("Actual  --->" + actualEsignCommentsUI)

    else:
        logger.error("Esign comment name  is not properly mentioned in UI")
        logger.error("Expected--->" + expectedEsignComments)
        logger.error("Actual  --->" + actualEsignCommentsUI)

    esignCommentQuery = "select sreason from auditaction where nauditcode=(select COUNT(*) from auditaction ) order by 1 desc"

    actualEsignCommentsDB = JDBC.retunOneValue(esignCommentQuery)

    if actualEsignCommentsDB == expectedEsignComments:
        logger.info("Esign comment is properly mentioned in db")
        logger.info("Expected--->" + expectedEsignComments)
        logger.info("Actual  --->" + actualEsignCommentsDB)
    else:
        logger.error("Esign comment is not properly mentioned db")
        logger.info("Expected--->" + expectedEsignComments)
        logger.info("Actual  --->" + actualEsignCommentsDB)

    if actualEsignCommentsDB == actualEsignCommentsUI:
        logger.info("Esign comment is same in the db and ui")
        logger.info("actualEsignCommentsDB--->" + actualEsignCommentsDB)
        logger.info("actualEsignCommentsUI  --->" + actualEsignCommentsUI)

    else:
        logger.error("Esign comment is not same in the db and ui")
        logger.error("actualEsignCommentsDB--->" + actualEsignCommentsDB)
        logger.error("actualEsignCommentsUI  --->" + actualEsignCommentsUI)



